from sqlite3 import connect
import requests
import json
from openpyxl import *
import pymysql

workbook=load_workbook("C:\\Users\\dev08\\Desktop\\파이썬 파일\\QAtextVer2.xlsx",data_only=True)
worksheet=workbook['조주영']
def excuteQuery(query):
    con=pymysql.connect(host="222.239.119.195",user="sbridge",password="wkddjrndl$#@!",db="pricegolf_re",charset="euckr",cursorclass=pymysql.cursors.DictCursor)
    cur=con.cursor()
    cur.execute(query)
    rows=cur.fetchall()
    con.close
    return rows

def connectApi(method,url,requestbody):
    token=str(requests.get('https://stg-api.pricegolf.co.kr/api/v1/getApiToken/sd2244').text)
    headers={'Content-Type' : 'application/json; chearset=utf-8', 'Authorization' : token}
    if method=='POST':
        try:
            if requestbody==None:
                requestbody=None
                res=requests.post(url,headers=headers)
            else:
                requestbody=json.loads(requestbody)
                res=requests.post(url,headers=headers,json=requestbody)
        except json.JSONDecodeError as e:
            return "클리이언트 오류","400"

    elif method=='GET':
        res=requests.get(url,headers=headers)
    elif method=='PUT':
        try:
            if requestbody==None:
                requestbody=None
                res=requests.put(url,headers=headers)
            else:
                requestbody=json.loads(requestbody)
                res=requests.put(url,headers=headers,json=requestbody)
        except json.JSONDecodeError as e:
            return "클리이언트 오류","400"
        
        
        
    elif method=='DELETE':
        res=requests.delete(url,headers=headers)
    try:
        res.raise_for_status()
        responseBody=res.text
        responseCode=res.status_code
        requestMethod=res.request.method
    except requests.exceptions.HTTPError as error:
        responseBody=error.response.text
        responseCode=res.status_code
        requestMethod=res.request.method

    return responseBody, responseCode
    #print(str(res.json()))


for i in range(2280,2286):
    url=worksheet.cell(i,11).value
    requestbody=worksheet.cell(i,12).value
    method=worksheet.cell(i,10).value
    expectResult=worksheet.cell(i,6).value
    print(i)
    print(url)

    responseBody, responseCode=connectApi(method,url,requestbody)
    if(expectResult==None or expectResult=="정상 값"):
        if( "200" in responseBody):
            if( "NONE" in responseBody):
                worksheet.cell(i,8).value="NONE"
            worksheet.cell(i,7).value="pass"
        else:
            worksheet.cell(i,7).value="fail"
        
    elif( expectResult in responseBody):
        worksheet.cell(i,7).value="pass"
    else:
        worksheet.cell(i,7).value="fail"
    worksheet.cell(i,13).value=responseCode
    worksheet.cell(i,14).value=responseBody

workbook.save('C:\\Users\\dev08\\Desktop\\파이썬 파일\\QAtextVer2.xlsx')
workbook.close()

import pandas as pd
path='C:\\Users\\dev\\Desktop\\'
files= ['2824713_result_(1)','2824713_result_(2)','2824713_result_(3)','2824713_result_(4)']
dfs=[]
for file in files:
    dfs.append(pd.read_excel(path+file+".xlsx"))    
productIdResult=[]
for df in dfs:
    productId=df.iloc[:,5].drop([0,1])
    productId=productId.str.split('-').str.get(0)
    productId=pd.to_numeric(productId)
    productIdResult.extend(productId.values)

coupangIds=set(productIdResult)

import sys, os,pymysql# '''mysql.connector, base64, requests,'''
from pandas import DataFrame
def connect_RDS(host,port,username,password,database):
    try:
        #con=mysql.connector.connect(host,user=username,passwd=password,db=database,port=port,use_unicode=True,charset='utf8')
        con=pymysql.connect(host=host,user=username,password=password,db=database,charset='utf8')
        cursor= con.cursor()
    except:
        #logging.error("RDS에 연결되지 않았습니다.")
        sys.exit(1)
    return con,cursor

port=3306
username='pr-web'
database='pricegolf_re'
password='WPdr(fk{d@}QtzN4VoVT'
host='pg-main.cluster-cijnjofh8bze.ap-northeast-2.rds.amazonaws.com'#:3306/'+database
#con=pymysql.connect(host=host,user=username,password=password,db=database,charset='utf8')

conn, cursor=connect_RDS(host,port,username,password,database)

# query="""select b.number,b.id,b.brand_name,a.coop_number,a.coop_stats from auction_product_coop a join auction_product b on a.product_number=b.number
# where b.id in ('1golf') and b.shipping_type and a.coop_name='lotteimall' and a.coop_stats='onsale' order by b.number,b.id desc"""
query="""select b.product_code, b.product_name, a.path from product_option_item a join auction_product b on a.product_number=b.number and a.oid=b.oid 
where b.product_stats=0 and shipping_type ='0'  and char_length(a.path) >22"""

cursor.execute(query)
data1=cursor.fetchall()
df6=DataFrame(data1)
conn.close()

dbCoupangIds=[]
dbCoupangIds.extend(pd.to_numeric(df6.iloc[:,0].values))
dbCoupangIdsSet=set(dbCoupangIds)
dbMscm=dbCoupangIdsSet-coupangIds
numsStr=str(dbMscm).replace("{","(").replace("}",")")
query="select product_number,coop_number from auction_product_coop where coop_name='coupang' and product_number in"
query+=numsStr
conn, cursor=connect_RDS(host,port,username,password,database)
cursor.execute(query)
data1=cursor.fetchall()
df1=DataFrame(data1)
path='C:\\Users\\dev\\Desktop\\'
df1.to_csv(path+"dbMscm.csv")
scmMdb=coupangIds-dbCoupangIdsSet
numStrRe=str(scmMdb).replace("{","(").replace("}",")")



